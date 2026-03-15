"""
Generate the "theBicycle" demo model Excel workbook.

Output: C:/allDev/ForecastAI2026.01/data/bicycle/bicycle_model.xlsx

Sheets
------
sites           — 31 sites (1 central WH + 5 country WH + 25 dealers, 5 per country)
items           — 22 racing bicycle parts (fast/medium/slow movers)
assets          — 5 complete racing bike models
bom             — multi-level Bill of Materials linking assets → sub-assemblies → parts
teams           — 6 teams (3 professional + 3 amateur)
team_assets     — which team uses which bike model and how many units
demand_actuals  — 3 years of weekly demand at each dealer site (≈ 120 k rows)
team_demand     — periodic maintenance/replacement demand driven by team schedules
"""

import os
import sys
from pathlib import Path
from datetime import date, timedelta
import math
import random

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# Output path
# ─────────────────────────────────────────────────────────────────────────────

OUT_DIR  = Path(__file__).resolve().parent.parent.parent / "data" / "bicycle"
OUT_FILE = OUT_DIR / "bicycle_model.xlsx"
OUT_DIR.mkdir(parents=True, exist_ok=True)

# Force UTF-8 output on Windows to avoid cp1252 encoding errors
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
elif sys.platform == "win32":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ─────────────────────────────────────────────────────────────────────────────
# Seed for reproducibility
# ─────────────────────────────────────────────────────────────────────────────
rng = np.random.default_rng(42)

# ─────────────────────────────────────────────────────────────────────────────
# MASTER DATA — SITES
# ─────────────────────────────────────────────────────────────────────────────

CENTRAL_WH = [
    # id, code, name, type, country, city, lat, lon, parent_site_id, lead_time_days
    (100, "CEN-BE", "Central Warehouse Brussels", "central_wh",
     "BE", "Brussels", 50.8503, 4.3517, None, 3),
]

COUNTRY_WH = [
    (201, "CW-FR", "France Country Warehouse",      "country_wh", "FR", "Paris",     48.8566,  2.3522, 100, 2),
    (202, "CW-DE", "Germany Country Warehouse",     "country_wh", "DE", "Munich",    48.1351, 11.5820, 100, 2),
    (203, "CW-ES", "Spain Country Warehouse",       "country_wh", "ES", "Madrid",    40.4168, -3.7038, 100, 3),
    (204, "CW-IT", "Italy Country Warehouse",       "country_wh", "IT", "Milan",     45.4654,  9.1859, 100, 2),
    (205, "CW-NL", "Netherlands Country Warehouse", "country_wh", "NL", "Amsterdam", 52.3676,  4.9041, 100, 1),
]

DEALERS = [
    # France
    (301, "FR-PAR", "Cycles Paris",      "dealer", "FR", "Paris",     48.8566,  2.3522, 201, 2),
    (302, "FR-LYN", "Vélo Lyon",         "dealer", "FR", "Lyon",      45.7640,  4.8357, 201, 2),
    (303, "FR-MRS", "Velo Marseille",    "dealer", "FR", "Marseille", 43.2965,  5.3698, 201, 3),
    (304, "FR-TLS", "Cycles Toulouse",   "dealer", "FR", "Toulouse",  43.6047,  1.4442, 201, 3),
    (305, "FR-BDX", "Bordeaux Bikes",    "dealer", "FR", "Bordeaux",  44.8378, -0.5792, 201, 3),
    # Germany
    (311, "DE-MUC", "München Rad",       "dealer", "DE", "Munich",    48.1351, 11.5820, 202, 1),
    (312, "DE-BER", "Berlin Cycles",     "dealer", "DE", "Berlin",    52.5200, 13.4050, 202, 2),
    (313, "DE-HAM", "Hamburg Fahrrad",   "dealer", "DE", "Hamburg",   53.5753, 10.0153, 202, 2),
    (314, "DE-CGN", "Köln Bikes",        "dealer", "DE", "Cologne",   50.9333,  6.9500, 202, 2),
    (315, "DE-STU", "Stuttgart Cycling", "dealer", "DE", "Stuttgart", 48.7758,  9.1829, 202, 2),
    # Spain
    (321, "ES-MAD", "Madrid Ciclismo",   "dealer", "ES", "Madrid",    40.4168, -3.7038, 203, 1),
    (322, "ES-BCN", "Barcelona Bici",    "dealer", "ES", "Barcelona", 41.3851,  2.1734, 203, 2),
    (323, "ES-VLC", "Valencia Ciclos",   "dealer", "ES", "Valencia",  39.4699, -0.3763, 203, 2),
    (324, "ES-SVQ", "Sevilla Ruedas",    "dealer", "ES", "Seville",   37.3891, -5.9845, 203, 3),
    (325, "ES-BIO", "Bilbao Bikes",      "dealer", "ES", "Bilbao",    43.2630, -2.9350, 203, 3),
    # Italy
    (331, "IT-MIL", "Milano Ciclismo",   "dealer", "IT", "Milan",     45.4654,  9.1859, 204, 1),
    (332, "IT-ROM", "Roma Bici",         "dealer", "IT", "Rome",      41.9028, 12.4964, 204, 2),
    (333, "IT-TOR", "Torino Cicli",      "dealer", "IT", "Turin",     45.0703,  7.6869, 204, 2),
    (334, "IT-FLR", "Firenze Bikes",     "dealer", "IT", "Florence",  43.7696, 11.2558, 204, 2),
    (335, "IT-VRN", "Verona Cycling",    "dealer", "IT", "Verona",    45.4384, 10.9916, 204, 2),
    # Netherlands
    (341, "NL-AMS", "Amsterdam Fietsen", "dealer", "NL", "Amsterdam", 52.3676,  4.9041, 205, 1),
    (342, "NL-RTD", "Rotterdam Fietsen", "dealer", "NL", "Rotterdam", 51.9225,  4.4792, 205, 1),
    (343, "NL-UTR", "Utrecht Cycles",    "dealer", "NL", "Utrecht",   52.0907,  5.1214, 205, 1),
    (344, "NL-EHV", "Eindhoven Bikes",   "dealer", "NL", "Eindhoven", 51.4416,  5.4697, 205, 1),
    (345, "NL-GRN", "Groningen Fietsen", "dealer", "NL", "Groningen", 53.2194,  6.5665, 205, 2),
]

ALL_SITES = CENTRAL_WH + COUNTRY_WH + DEALERS
SITE_COLS = ["site_id","site_code","site_name","site_type","country","city",
             "latitude","longitude","parent_site_id","lead_time_days"]

# ─────────────────────────────────────────────────────────────────────────────
# MASTER DATA — ITEMS
# ─────────────────────────────────────────────────────────────────────────────

# (id, code, name, category, description, unit, weight_g, price_eur, mover_class, base_lambda_week)
# base_lambda_week = mean weekly demand units at an average dealer
ITEMS = [
    # Consumables — fast movers
    (1,  "LB-FI",  "Chain Lube Finish Line 500ml",               "Consumable",  "Dry/wet chain lubricant 500 ml bottle",     "bottle",  520,   12,  "A", 8.0),
    (2,  "TU-CT",  "Inner Tube Continental 700×28",              "Wheels",      "Butyl inner tube, Presta valve",            "each",     90,    8,  "A", 6.0),
    (3,  "TI-CT",  "Tire Continental Grand Prix 5000 S TR 28mm", "Wheels",      "TL-ready road tire, 28 mm",                 "each",    225,   48,  "A", 4.0),
    (4,  "CH-DA",  "Chain Shimano Dura-Ace CN-M9100 12sp",       "Drivetrain",  "12-speed HG-EV chain, 116 links",           "each",    230,   65,  "A", 2.5),
    (5,  "BP-DA",  "Brake Pads Dura-Ace R9200 (4 pcs)",          "Braking",     "Carbon/resin pads for R9200 caliper",       "set",      48,   28,  "A", 2.0),
    # Medium movers
    (6,  "CS-DA",  "Cassette Shimano DA R9200 12sp 11-30T",      "Drivetrain",  "12-speed HG-EV, titanium sprockets",        "each",    181,  185,  "B", 1.2),
    (7,  "BC-DA",  "Brake Caliper Set Shimano DA R9200",         "Braking",     "Hydraulic disc caliper set front+rear",     "set",     320,  320,  "B", 0.8),
    (8,  "SD-FZ",  "Saddle Fizik Ares R1 Carbon",                "Cockpit",     "Carbon-railed road saddle, 140 mm",         "each",    156,  280,  "B", 0.6),
    (9,  "HB-RC",  "Handlebar Ritchey WCS Carbon 42 cm",         "Cockpit",     "Compact drop road bar, -6° flare",          "each",    185,  145,  "B", 0.5),
    (10, "ST-RC",  "Stem Ritchey WCS Carbon 100 mm",             "Cockpit",     "Carbon road stem ±6°",                      "each",    110,  110,  "B", 0.5),
    # Slow movers
    (11, "SP-EL",  "Seatpost Thomson Elite 27.2×350 mm",         "Cockpit",     "Alloy offset seatpost",                     "each",    220,  135,  "C", 0.4),
    (12, "HC-GR",  "GPS Garmin Edge 1040 Solar",                 "Electronics", "Solar-assist GPS cycling computer",         "each",    133,  599,  "C", 0.4),
    (13, "CK-DA",  "Crankset Shimano DA FC-R9200 52/36T",        "Drivetrain",  "12sp hollow-tech II, 170 mm",               "each",    614,  590,  "C", 0.3),
    (14, "RD-DA",  "Rear Derailleur Shimano DA RD-R9200",        "Drivetrain",  "12sp wireless-ready rear derailleur",       "each",    198,  420,  "C", 0.3),
    (15, "FD-DA",  "Front Derailleur Shimano DA FD-R9200",       "Drivetrain",  "12sp braze-on front derailleur",            "each",    104,  185,  "C", 0.2),
    (16, "PM-GR",  "Power Meter Garmin Rally RS100",             "Electronics", "Single-sided SPD-SL power meter pedals",   "pair",    374,  699,  "C", 0.15),
    (17, "BT-DI",  "Di2 Battery BT-DN300 + Charger Kit",        "Electronics", "Internal Li-ion battery + SM-BCR2 charger","kit",      55,   95,  "C", 0.15),
    # Very slow movers
    (18, "WS-ZP",  "Wheelset Zipp 303 Firecrest TL Disc",       "Wheels",      "Carbon clincher/TL 45 mm, Centre Lock",     "pair",   1490, 1850,  "C", 0.12),
    (19, "FK-CN",  "Carbon Road Fork (Canyon/Enve compatible)", "Frame",       "Tapered 1-1/8–1.5″ carbon road fork",       "each",    320,  380,  "C", 0.08),
    (20, "FR-RD",  "Carbon Road Frame 56 cm",                   "Frame",       "Monocoque carbon road frame, BB86",         "each",   1050, 1200,  "C", 0.06),
    # Sub-assembly virtual items (used in BOM, not ordered individually)
    (21, "SA-DRV", "Drivetrain Service Kit",                    "Sub-assembly","Cassette + Chain — annual service bundle",  "kit",     411,  250,  "B", 0.8),
    (22, "SA-BRK", "Brake Service Kit",                         "Sub-assembly","Pads (×2 sets) + bleed kit — annual svc",  "kit",      96,   56,  "B", 0.5),
]
ITEM_COLS = ["item_id","item_code","item_name","category","description",
             "unit","weight_g","price_eur","mover_class","base_lambda_week"]

# lookup dicts
ITEM_BY_CODE = {row[1]: row for row in ITEMS}
ITEM_ID_BY_CODE = {row[1]: row[0] for row in ITEMS}

# ─────────────────────────────────────────────────────────────────────────────
# MASTER DATA — ASSETS (complete bike models)
# ─────────────────────────────────────────────────────────────────────────────

ASSETS = [
    # id, code, name, type, brand, model_year, description, image_url
    (1, "CAN-ULT", "Canyon Ultimate CFR Disc",
     "Road - Climbing",   "Canyon",       2025,
     "Lightweight all-rounder, 6.7 kg UCI legal; flagship carbon monocoque frame",
     "https://www.canyon.com/dw/image/v2/BCML_PRD/on/demandware.static/-/Sites/default/dw6893c745/images/generated/highRes/canyon-ultimate-cfr-disc-2025.jpg"),

    (2, "CAN-AER", "Canyon Aeroad CFR Di2",
     "Road - Aero",        "Canyon",       2025,
     "Canyon's most aerodynamic road bike; integrated cockpit, wind-cheating tube shapes",
     "https://www.canyon.com/dw/image/v2/BCML_PRD/on/demandware.static/-/Sites/default/dw1234/images/generated/highRes/canyon-aeroad-cfr-di2-2025.jpg"),

    (3, "SPZ-TAR", "Specialized Tarmac SL8 Pro",
     "Road - All-round",   "Specialized",  2025,
     "Balanced performance: 6.8 kg frame-set, Tarmac geometry, FACT 12r carbon",
     "https://assets.specialized.com/i/specialized/95324-100_TARMAC-SL8-PRO-ULTEGRA-DI2_BLK-CLGRY_HERO.jpg"),

    (4, "TRK-MAD", "Trek Madone SLR 9",
     "Road - Aero",        "Trek",         2025,
     "IsoFlow technology for comfort + speed; integrated bar/stem, KVF tube profiles",
     "https://trek.scene7.com/is/image/TrekBicycleProducts/MadoneSLR9_23_36282_A_Primary"),

    (5, "CER-P5",  "Cervelo P5 Disc",
     "TT/Triathlon",       "Cervelo",      2025,
     "Full UCI-legal TT bike; Aerozone cockpit, IQ2 integrated headset, disc brakes",
     "https://www.cervelo.com/en/bikes/p-series/p5"),
]
ASSET_COLS = ["asset_id","asset_code","asset_name","asset_type","brand",
              "model_year","description","image_url"]

# ─────────────────────────────────────────────────────────────────────────────
# BOM — multi-level
# ─────────────────────────────────────────────────────────────────────────────
# Columns: bom_id, asset_id, level, parent_item_id (NULL=direct child of asset),
#          child_item_id, quantity, unit, notes

BOM_ROWS = []
_bom_id = 1

def _bom(asset_id, level, parent_item_id, child_code, qty, unit="each", notes=""):
    global _bom_id
    child_id = ITEM_ID_BY_CODE[child_code]
    BOM_ROWS.append((_bom_id, asset_id, level, parent_item_id, child_id, qty, unit, notes))
    _bom_id += 1

# ─── Level-1 direct components for each bike ─────────────────────────────────
# All 5 bikes share mostly identical BOMs (minor differences noted)
for _asset_id in [1, 2, 3, 4, 5]:  # all bikes
    _bom(_asset_id, 1, None, "FR-RD",  1, notes="Frame 56 cm")
    _bom(_asset_id, 1, None, "FK-CN",  1, notes="Carbon road fork")
    _bom(_asset_id, 1, None, "CK-DA",  1, notes="Crankset 52/36 T")
    _bom(_asset_id, 1, None, "SA-DRV", 1, unit="kit",
         notes="Drivetrain Service Kit — contains sub-items below")
    _bom(_asset_id, 1, None, "BC-DA",  1, unit="set", notes="Brake caliper set")
    _bom(_asset_id, 1, None, "SA-BRK", 1, unit="kit",
         notes="Brake Service Kit — contains sub-items below")
    _bom(_asset_id, 1, None, "WS-ZP",  1, unit="pair", notes="Wheelset")
    _bom(_asset_id, 1, None, "TI-CT",  2, notes="Tires (front + rear)")
    _bom(_asset_id, 1, None, "TU-CT",  2, notes="Inner tubes (front + rear)")
    _bom(_asset_id, 1, None, "SD-FZ",  1, notes="Saddle")
    _bom(_asset_id, 1, None, "PM-GR",  1, unit="pair", notes="Power meter pedals")
    _bom(_asset_id, 1, None, "HC-GR",  1, notes="GPS head unit")
    _bom(_asset_id, 1, None, "BT-DI",  1, unit="kit", notes="Di2 battery + charger")

# Road bikes (not TT): standard cockpit
for _asset_id in [1, 2, 3, 4]:
    _bom(_asset_id, 1, None, "HB-RC",  1, notes="Road handlebar")
    _bom(_asset_id, 1, None, "ST-RC",  1, notes="Road stem")
    _bom(_asset_id, 1, None, "SP-EL",  1, notes="Seatpost")

# TT bike (CER-P5): aero cockpit (reuse same items, different note)
_bom(5, 1, None, "HB-RC", 1, notes="Aero TT base bar (road bar equivalent)")
_bom(5, 1, None, "ST-RC", 1, notes="TT stem adapter")
_bom(5, 1, None, "SP-EL", 1, notes="TT seatpost")

# ─── Level-2: sub-assembly children ──────────────────────────────────────────
SA_DRV_ID = ITEM_ID_BY_CODE["SA-DRV"]
SA_BRK_ID = ITEM_ID_BY_CODE["SA-BRK"]

for _asset_id in [1, 2, 3, 4, 5]:
    # SA-DRV expands to: RD + FD + cassette + chain
    _bom(_asset_id, 2, SA_DRV_ID, "RD-DA",  1)
    _bom(_asset_id, 2, SA_DRV_ID, "FD-DA",  1)
    _bom(_asset_id, 2, SA_DRV_ID, "CS-DA",  1)
    _bom(_asset_id, 2, SA_DRV_ID, "CH-DA",  1)
    # SA-BRK expands to: 2 sets of brake pads + lube
    _bom(_asset_id, 2, SA_BRK_ID, "BP-DA",  2, unit="set")
    _bom(_asset_id, 2, SA_BRK_ID, "LB-FI",  1, unit="bottle", notes="Brake bleed fluid")

BOM_COLS = ["bom_id","asset_id","level","parent_item_id","child_item_id",
            "quantity","unit","notes"]

# ─────────────────────────────────────────────────────────────────────────────
# MASTER DATA — TEAMS
# ─────────────────────────────────────────────────────────────────────────────

TEAMS = [
    # id, code, name, type, country, contact_email, num_riders, num_bikes
    (1, "UAE",  "UAE Team Emirates",        "professional", "AE",
     "logistics@uae-teamemirates.com",  30, 90),
    (2, "VLB",  "Visma-Lease a Bike",       "professional", "NL",
     "service@vismabike.com",           30, 90),
    (3, "INE",  "Ineos Grenadiers",         "professional", "GB",
     "race.ops@ineos-grenadiers.com",   30, 90),
    (4, "TDF",  "Équipe Amateurs Tour",     "amateur",      "FR",
     "contact@equipe-amateurs.fr",      12, 15),
    (5, "BCC",  "Berlin Cycling Club",      "amateur",      "DE",
     "info@berlin-cc.de",               15, 18),
    (6, "BCN",  "Barcelona Riders Club",    "amateur",      "ES",
     "riders@barcelonacc.es",           10, 12),
]
TEAM_COLS = ["team_id","team_code","team_name","team_type","country",
             "contact_email","num_riders","num_bikes"]

# ─────────────────────────────────────────────────────────────────────────────
# TEAM ASSETS
# ─────────────────────────────────────────────────────────────────────────────

TEAM_ASSETS = [
    # id, team_id, asset_id, quantity, season, notes
    (1,  1, 2, 45, 2025, "UAE - Aeroad for flat/sprint stages"),
    (2,  1, 3, 45, 2025, "UAE - Tarmac for mountain stages"),
    (3,  2, 1, 50, 2025, "Visma - Ultimate CFR for all terrain"),
    (4,  2, 2, 40, 2025, "Visma - Aeroad for TTs"),
    (5,  3, 3, 45, 2025, "Ineos - Tarmac SL8 general use"),
    (6,  3, 4, 45, 2025, "Ineos - Madone for aerodynamic stages"),
    (7,  4, 1,  8, 2025, "TDF Amateurs - Ultimate CFR"),
    (8,  4, 3,  7, 2025, "TDF Amateurs - Tarmac SL8"),
    (9,  5, 2, 18, 2025, "Berlin CC - Aeroad"),
    (10, 6, 3, 12, 2025, "Barcelona Riders - Tarmac SL8"),
]
TA_COLS = ["id","team_id","asset_id","quantity","season","notes"]

# ─────────────────────────────────────────────────────────────────────────────
# DEMAND GENERATION
# ─────────────────────────────────────────────────────────────────────────────

# Date range: 3 full calendar years of weekly data
START_DATE = date(2022, 1,  3)   # first Monday of 2022
END_DATE   = date(2024, 12, 30)  # last Monday of 2024

# Build weekly date index (Mondays)
weekly_dates = []
d = START_DATE
while d <= END_DATE:
    weekly_dates.append(d)
    d += timedelta(weeks=1)
N_WEEKS = len(weekly_dates)

# Country demand multipliers (cycling culture weighting)
COUNTRY_MULT = {"NL": 1.40, "FR": 1.20, "DE": 1.10, "IT": 1.00, "ES": 0.90}

# Seasonal multiplier by ISO week number (1-52)
def seasonal(week_num: int) -> float:
    """Return a seasonal demand multiplier for a given ISO week (1-52)."""
    # Approximated with a cosine: peak week ~26 (late June)
    angle = 2 * math.pi * (week_num - 26) / 52
    base = 1.0 + 0.7 * math.cos(angle + math.pi)   # 0.3 at min, 1.7 at max
    return max(0.1, base)

# Annual trend: +20% over 3 years (cycling boom post-2020)
def trend(week_index: int, total_weeks: int) -> float:
    return 1.0 + 0.20 * (week_index / total_weeks)

# Items for which we generate demand: only real items (not virtual sub-assemblies)
DEMAND_ITEMS = [row for row in ITEMS if row[0] <= 20]   # exclude virtual SA items

print(f"Generating demand: {len(DEMAND_ITEMS)} items × {len(DEALERS)} dealers × {N_WEEKS} weeks …")

demand_rows = []   # (unique_id, item_id, site_id, channel, date, qty)

for item in DEMAND_ITEMS:
    item_id, item_code, item_name = item[0], item[1], item[2]
    base_lam = item[9]   # mean weekly demand at average dealer

    for site_row in DEALERS:
        site_id, site_code, site_name = site_row[0], site_row[1], site_row[2]
        country = site_row[4]
        country_m = COUNTRY_MULT.get(country, 1.0)
        # Slight per-site variance to create realistic variety
        site_factor = rng.uniform(0.7, 1.4)
        unique_id = f"{item_id}_{site_id}"

        for wi, dt in enumerate(weekly_dates):
            week_num = dt.isocalendar()[1]
            lam = base_lam * country_m * site_factor * seasonal(week_num) * trend(wi, N_WEEKS)
            # Use Poisson for discrete demand; very slow items: may be 0 most weeks
            qty = int(rng.poisson(max(lam, 0.01)))
            if qty > 0:
                demand_rows.append((unique_id, item_id, site_id, "", str(dt), qty))

print(f"  → {len(demand_rows):,} non-zero demand rows generated")

DEMAND_COLS = ["unique_id","item_id","site_id","channel","date","qty"]

# ─────────────────────────────────────────────────────────────────────────────
# TEAM DEMAND (causal — maintenance events)
# ─────────────────────────────────────────────────────────────────────────────
# Professional teams: race season March–October (31 weeks)
# Amateur teams: riding season April–September (26 weeks)
# Maintenance intervals (pro): chain every 4 wks, cassette every 8 wks, tires every 6 wks
# Maintenance intervals (amateur): chain every 6 wks, cassette every 12 wks, tires every 8 wks

TEAM_DEMAND_COLS = ["id","team_id","item_id","asset_id","date","qty","event_type","notes"]
team_demand_rows = []
_td_id = 1

PRO_SEASON_WEEKS = [w for w in weekly_dates
                    if 9 <= w.isocalendar()[1] <= 43]   # Mar–Oct
AMT_SEASON_WEEKS = [w for w in weekly_dates
                    if 14 <= w.isocalendar()[1] <= 39]  # Apr–Sep

# Map team → asset quantities
team_assets_map = {}
for row in TEAM_ASSETS:
    _, tid, aid, qty, _, _ = row
    team_assets_map.setdefault(tid, {})[aid] = qty

def _add_team_demand(team_id, asset_id, item_code, qty_per_bike,
                     season_weeks, interval_weeks, event_type, note):
    global _td_id
    n_bikes = team_assets_map.get(team_id, {}).get(asset_id, 0)
    if n_bikes == 0:
        return
    item_id = ITEM_ID_BY_CODE.get(item_code)
    if item_id is None:
        return
    for wi, wdt in enumerate(season_weeks):
        if wi % interval_weeks == 0:
            qty = n_bikes * qty_per_bike
            team_demand_rows.append((
                _td_id, team_id, item_id, asset_id,
                str(wdt), qty, event_type, note,
            ))
            _td_id += 1

# Professional teams (teams 1-3)
for team_id in [1, 2, 3]:
    for asset_id in team_assets_map.get(team_id, {}):
        _add_team_demand(team_id, asset_id, "CH-DA",  1, PRO_SEASON_WEEKS, 4,  "maintenance", "Chain replacement")
        _add_team_demand(team_id, asset_id, "CS-DA",  1, PRO_SEASON_WEEKS, 8,  "maintenance", "Cassette replacement")
        _add_team_demand(team_id, asset_id, "TI-CT",  2, PRO_SEASON_WEEKS, 6,  "maintenance", "Tire replacement (pair)")
        _add_team_demand(team_id, asset_id, "BP-DA",  1, PRO_SEASON_WEEKS, 10, "maintenance", "Brake pad replacement")
        _add_team_demand(team_id, asset_id, "LB-FI",  1, PRO_SEASON_WEEKS, 2,  "maintenance", "Chain lubrication")

# Amateur teams (teams 4-6)
for team_id in [4, 5, 6]:
    for asset_id in team_assets_map.get(team_id, {}):
        _add_team_demand(team_id, asset_id, "CH-DA",  1, AMT_SEASON_WEEKS, 6,  "maintenance", "Chain replacement")
        _add_team_demand(team_id, asset_id, "CS-DA",  1, AMT_SEASON_WEEKS, 12, "maintenance", "Cassette replacement")
        _add_team_demand(team_id, asset_id, "TI-CT",  2, AMT_SEASON_WEEKS, 8,  "maintenance", "Tire replacement")
        _add_team_demand(team_id, asset_id, "LB-FI",  1, AMT_SEASON_WEEKS, 3,  "maintenance", "Chain lubrication")

print(f"  → {len(team_demand_rows):,} team demand event rows generated")

# ─────────────────────────────────────────────────────────────────────────────
# WRITE EXCEL FILE
# ─────────────────────────────────────────────────────────────────────────────

HDR_FILL  = PatternFill("solid", fgColor="1F3864")   # dark navy
HDR_FONT  = Font(bold=True, color="FFFFFF")
ALT_FILL  = PatternFill("solid", fgColor="EEF2FF")   # light blue-grey
NO_FILL   = PatternFill("none")
THIN      = Side(style="thin", color="CCCCCC")
BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def _write_sheet(wb, title: str, columns: list, rows: list):
    ws = wb.create_sheet(title=title)
    # Header row
    for ci, col in enumerate(columns, 1):
        cell = ws.cell(1, ci, col)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    ws.row_dimensions[1].height = 20

    # Data rows
    for ri, row_data in enumerate(rows, 2):
        fill = ALT_FILL if ri % 2 == 0 else NO_FILL
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(ri, ci, val)
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")

    # Auto-fit column widths (approx)
    for ci, col in enumerate(columns, 1):
        col_letter = get_column_letter(ci)
        max_len = len(col)
        for row_data in rows[:500]:   # sample first 500 rows for speed
            v = str(row_data[ci - 1]) if row_data[ci - 1] is not None else ""
            max_len = max(max_len, min(len(v), 60))
        ws.column_dimensions[col_letter].width = max_len + 3

    # Freeze header
    ws.freeze_panes = "A2"
    return ws


print("Building Excel workbook …")
wb = Workbook()
wb.remove(wb.active)   # remove default blank sheet

_write_sheet(wb, "sites",          SITE_COLS,    ALL_SITES)
_write_sheet(wb, "items",          ITEM_COLS,    ITEMS)
_write_sheet(wb, "assets",         ASSET_COLS,   ASSETS)
_write_sheet(wb, "bom",            BOM_COLS,     BOM_ROWS)
_write_sheet(wb, "teams",          TEAM_COLS,    TEAMS)
_write_sheet(wb, "team_assets",    TA_COLS,      TEAM_ASSETS)
_write_sheet(wb, "team_demand",    TEAM_DEMAND_COLS, team_demand_rows)

# demand_actuals — large, write in chunks for memory efficiency
print(f"Writing demand_actuals sheet ({len(demand_rows):,} rows) …")
ws_demand = wb.create_sheet(title="demand_actuals")
for ci, col in enumerate(DEMAND_COLS, 1):
    cell = ws_demand.cell(1, ci, col)
    cell.font = HDR_FONT
    cell.fill = HDR_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER
ws_demand.row_dimensions[1].height = 20
ws_demand.freeze_panes = "A2"

for ri, row_data in enumerate(demand_rows, 2):
    for ci, val in enumerate(row_data, 1):
        ws_demand.cell(ri, ci, val)

# Set column widths for demand sheet
col_widths = [22, 8, 8, 8, 12, 8]
for ci, w in enumerate(col_widths, 1):
    ws_demand.column_dimensions[get_column_letter(ci)].width = w

wb.save(OUT_FILE)
print(f"\n✓ Saved: {OUT_FILE}")
print(f"  Sheets: {[s.title for s in wb.worksheets]}")
print(f"  Sites          : {len(ALL_SITES):>6,}")
print(f"  Items          : {len(ITEMS):>6,}")
print(f"  Assets         : {len(ASSETS):>6,}")
print(f"  BOM rows       : {len(BOM_ROWS):>6,}")
print(f"  Teams          : {len(TEAMS):>6,}")
print(f"  Team assets    : {len(TEAM_ASSETS):>6,}")
print(f"  Team demand    : {len(team_demand_rows):>6,}")
print(f"  Demand actuals : {len(demand_rows):>6,}")
